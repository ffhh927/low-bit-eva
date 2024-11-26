import torch
import torchvision
import torchvision.transforms as transforms
from kfac import EVA, EVA8bit_Fused, EVA8bit, EVA_Fused, EVA8bit_CPU
from bitsandbytes.optim import Adam8bit, SGD8bit, Adam32bit, SGD32bit 
from shampoo import ShampooSGD4bit_svd, ShampooSGD4bit
import torch_optimizer as optim
from model.resnet20 import ResNet as resnet20
from model.vit import ViT
from model.autoencoder import Autoencoder
import xlsxwriter
from openpyxl import Workbook, load_workbook
import torch.nn as nn
import os
# Initialize Excel workbook and worksheet for logging results
# workbook = xlsxwriter.Workbook('./logs/cifar10_iteration2000_training_times1.xlsx')
file_path = './logs/cifar10_iteration2000_training_times_NVIDIA_RTX_A6000.xlsx'
# sheet_name = 'test log.'

from datetime import datetime
sheet_name = datetime.now().strftime("%Y-%m-%d %H-%M-%S")


try:
    # 尝试加载现有文件
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
    else:
        # 文件不存在，创建新的工作簿
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
except Exception as e:
    print(f"Error loading workbook: {e}")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

# 检查表头是否已写入
if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
    sheet.delete_rows(1)  # 确保清理第一行的空单元格
    sheet.append(["Model", "Optimizer", "Batch Size", "Avg Forward Time (ms)", 
                  "Avg Backward Time (ms)", "Avg Optimize Time (ms)", "Avg Step Time (ms)"])

# 保存文件
workbook.save(file_path)

# Function to initialize and configure the CIFAR-10 dataset
def get_cifar10_data(batch_size):
    transform = transforms.Compose([
        transforms.RandomCrop(32, padding=4),
        transforms.Resize(32),
        transforms.RandomHorizontalFlip(),
        transforms.ToTensor(),
        transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010)),
    ])
    trainset = torchvision.datasets.CIFAR10(root='./data/cifar10', train=True, download=True, transform=transform)
    trainloader = torch.utils.data.DataLoader(trainset, batch_size=batch_size, shuffle=True, num_workers=2)
    return trainloader

# Main training and timing function
def run_training(net, optimizer, criterion, trainloader):
    forward_start, forward_end = torch.cuda.Event(enable_timing=True), torch.cuda.Event(enable_timing=True)
    backward_start, backward_end = torch.cuda.Event(enable_timing=True), torch.cuda.Event(enable_timing=True)
    optimize_start, optimize_end = torch.cuda.Event(enable_timing=True), torch.cuda.Event(enable_timing=True)
    step_start, step_end = torch.cuda.Event(enable_timing=True), torch.cuda.Event(enable_timing=True)
    
    sum_forward_time, sum_backward_time, sum_optimize_time, sum_step_time, num = 0, 0, 0, 0, 0
    warmup = 20
    iteration = 200
    
    for epoch in range(1000):
        for i, data in enumerate(trainloader, 0):
            step_start.record()
            inputs, labels = data[0].to(device), data[1].to(device)
            forward_start.record()
            optimizer.zero_grad()
            outputs = net(inputs)
            loss = criterion(outputs, labels)
            torch.cuda.synchronize()
            forward_end.record()

            backward_start.record()
            loss.backward()
            torch.cuda.synchronize()
            backward_end.record()

            optimize_start.record()
            optimizer.step()
            torch.cuda.synchronize()
            optimize_end.record()
            step_end.record()

            if num >= warmup:
                torch.cuda.synchronize()  # 同步
                sum_forward_time += forward_start.elapsed_time(forward_end)
                sum_backward_time += backward_start.elapsed_time(backward_end)
                sum_optimize_time += optimize_start.elapsed_time(optimize_end)
                sum_step_time += step_start.elapsed_time(step_end)
            num += 1
            if num - warmup >= iteration:
                avg_forward_time = sum_forward_time / (num - warmup) 
                avg_backward_time = sum_backward_time / (num - warmup) 
                avg_optimize_time = sum_optimize_time / (num - warmup) 
                avg_step_time = sum_step_time / (num - warmup) 
                print(num)
                print(epoch)

                return avg_forward_time, avg_backward_time, avg_optimize_time, avg_step_time

    avg_forward_time = sum_forward_time / (num - warmup) 
    avg_backward_time = sum_backward_time / (num - warmup) 
    avg_optimize_time = sum_optimize_time / (num - warmup) 
    avg_step_time = sum_step_time / (num - warmup) 
    print(num)
    print(epoch)
    return avg_forward_time, avg_backward_time, avg_optimize_time, avg_step_time

# Function to run training and log results for different batch sizes and optimizers
# def run_training_with_optimizer(model_name, optimizer_name,  criterion, batch_sizes):
#     device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
#     model = select_model(model_name).to(device)
#     optimizer = select_optimizer(model, optimizer_name)
#     worksheet_name = model_name + '_' + optimizer_name
#     worksheet = workbook.add_worksheet(f"{worksheet_name} logs")
#     worksheet.write_row(0, 0, ["Batch Size", "Avg Forward Time (ms)", "Avg Backward Time (ms)", "Avg Optimize Time (ms)", "Avg Step Time (ms)"])
    
#     row = 1
#     for batch_size in batch_sizes:
#         trainloader = get_cifar10_data(batch_size)
#         avg_forward, avg_backward, avg_optimize, avg_step = run_training(model, optimizer, criterion, trainloader)
#         worksheet.write_row(row, 0, [batch_size, avg_forward, avg_backward, avg_optimize, avg_step])
#         print(f"{worksheet_name} Batch Size: {batch_size}, Forward: {avg_forward:.4f}ms, Backward: {avg_backward:.4f}ms, Optimize: {avg_optimize:.4f}ms, Step: {avg_step:.4f}ms")
#         row += 1

def run_training_with_optimizer(sheet, model_name, optimizer_name, criterion, batch_sizes):
    device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
    model = select_model(model_name).to(device)
    optimizer = select_optimizer(model, optimizer_name)  
    
    # 从最后一行开始追加
    for batch_size in batch_sizes:
        trainloader = get_cifar10_data(batch_size)
        avg_forward, avg_backward, avg_optimize, avg_step = run_training(model, optimizer, criterion, trainloader)
        sheet.append([model_name, optimizer_name, batch_size, avg_forward, avg_backward, avg_optimize, avg_step])
        print(f"Model: {model_name} Optimizer: {optimizer_name} Batch Size: {batch_size}, "
              f"Forward: {avg_forward:.4f}ms, Backward: {avg_backward:.4f}ms, "
              f"Optimize: {avg_optimize:.4f}ms, Step: {avg_step:.4f}ms")
        
    workbook.save(file_path)

def select_optimizer(net, optimizer_choice="adam32bit", lr=2e-5):
        # 根据命令行选择优化器
    if optimizer_choice == 'eva':
        optimizer = EVA(net, lr=lr, sgd_momentum=0.9)
    elif optimizer_choice == 'eva8bit':
        optimizer = EVA8bit(net, lr=lr, sgd_momentum=0.9)
    elif optimizer_choice == 'eva8bit_cpu':
        optimizer = EVA8bit_CPU(net, lr=lr, sgd_momentum=0.9)
    elif optimizer_choice == 'eva_fused':
        optimizer = EVA_Fused(net, lr=lr, sgd_momentum=0.9)
    elif optimizer_choice == 'eva8bit_fused':
        optimizer = EVA8bit_Fused(net, lr=lr, sgd_momentum=0.9)
    elif optimizer_choice == 'adam8bit':
        optimizer = Adam8bit(net.parameters(), lr=lr)          
    elif optimizer_choice == 'adam32bit':
        optimizer = Adam32bit(net.parameters(), lr=lr)  
    elif optimizer_choice == 'sgd8bit':
        optimizer = SGD8bit(net.parameters(), lr=lr, momentum=0.9) 
    elif optimizer_choice == 'sgd32bit':
        optimizer = SGD32bit(net.parameters(), lr=lr, momentum=0.9)
    elif optimizer_choice == 'shampoo32bit':
        optimizer = optim.Shampoo(net.parameters(), lr=0.01, momentum=0.9, weight_decay=0.0005)
    elif optimizer_choice == 'shampoo4bit_svd':
        optimizer = ShampooSGD4bit_svd(net.parameters(),
                        lr=0.1,
                        momentum=0.9,
                        weight_decay=0.0005,
                        nesterov=False,
                        start_prec_step=1,
                        stat_compute_steps=100,
                        prec_compute_steps=500,
                        stat_decay=0.95,
                        matrix_eps=1e-6,
                        prec_maxorder=1200,
                        prec_bits=4,
                        min_lowbit_size=4096,
                        quan_blocksize=64)   
    elif optimizer_choice == 'shampoo4bit':
        optimizer = ShampooSGD4bit(net.parameters(),
                        lr=0.1,
                        momentum=0.9,
                        weight_decay=0.0005,
                        nesterov=False,
                        start_prec_step=1,
                        stat_compute_steps=100,
                        prec_compute_steps=500,
                        stat_decay=0.95,
                        matrix_eps=1e-6,
                        prec_maxorder=1200,
                        prec_bits=4,
                        min_lowbit_size=4096,
                        quan_blocksize=64)   
    else:
        raise ValueError("Unsupported optimizer name. Choose 'eva','eva8bit','eva_fused','eva8bit_fused','adam8it', \
                         adam32bit', 'sgd32bit', 'sgd8bit', 'shampoo4bit', 'shampoo4bit_svd', 'shampoo32bit'.")
    return optimizer


def select_model(model_name):
    if model_name=='vit':
        net = ViT(
            image_size = 32,
            patch_size = 4,
            num_classes = 10,
            dim = 512,
            depth = 6,
            heads = 8,
            mlp_dim = 512,
            dropout = 0.1,
            emb_dropout = 0.1
        )
    elif model_name == 'res20':
        net = resnet20()
    elif model_name == 'autoencoder':
        net = Autoencoder()
    else:
        raise ValueError("Unsupported model name. Choose 'vit','res20','autoencoder'.")
    return net
    
# Running the experiments
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
criterion = nn.CrossEntropyLoss()
# batch_sizes = [2**i for i in range(10)]  # Batch sizes 1, 2, 4, ..., 512
# batch_sizes = [1, 8, 32, 128, 512]  # Batch sizes 1, 8, 32, 128, 512
batch_sizes = [1]  # Batch sizes 1, 8, 32, 128, 512

# Example usage with different optimizers
# 定义模型和优化器的名称列表
models = ['res20', 'autoencoder', 'vit']
# optimizers = [
#     'eva', 'eva8bit', 'eva8bit_cpu', 
#     'adam32bit', 'adam8bit', 
#     'sgd32bit', 'sgd8bit',
#     'shampoo4bit', 'shampoo4bit_svd'
# ]
optimizers = [
    'eva8bit_cpu',
]

# 使用嵌套循环来简化调用
for model_name in models:
    for optimizer_name in optimizers:
        run_training_with_optimizer(sheet, model_name, optimizer_name, criterion, batch_sizes)


# Close the workbook after logging all results
workbook.close()
